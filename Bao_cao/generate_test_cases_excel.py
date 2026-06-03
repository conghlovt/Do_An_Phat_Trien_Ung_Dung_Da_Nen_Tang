from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_FILE = "Kich_Ban_Kiem_Thu_Chi_Tiet.xlsx"
SUMMARY_SHEET = "Tổng Quan (Summary)"
DETAIL_SHEET = "Danh Sách Test Cases"

NAVY = "1F4E78"
WHITE = "FFFFFF"
ZEBRA = "F2F5F8"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "E2F0D9"
BORDER = "D9E2EC"
TEXT_DARK = "1F2937"


def numbered(items: list[str]) -> str:
    return "\n".join(f"{index}. {item}" for index, item in enumerate(items, 1))


def build_test_cases() -> list[dict[str, str]]:
    groups: list[tuple[str, str, list[tuple[str, str, list[str], list[str]]]]] = [
        (
            "AUTH",
            "Xác thực & Phiên",
            [
                (
                    "Đăng ký tài khoản Customer hợp lệ",
                    "Người dùng chưa có tài khoản với email dùng để đăng ký.",
                    [
                        "Mở màn hình Register.",
                        "Nhập username tối thiểu 2 ký tự, email hợp lệ, password và confirm password từ 6 ký tự trở lên.",
                        "Chọn Account Type là Customer và nhấn Create Account.",
                    ],
                    [
                        "Hệ thống tạo tài khoản role customer thành công.",
                        "Người dùng được điều hướng về màn hình Login.",
                        "Không hiển thị lỗi validation hoặc lỗi API.",
                    ],
                ),
                (
                    "Đăng ký tài khoản Partner hợp lệ",
                    "Người dùng chưa có tài khoản với email dùng để đăng ký.",
                    [
                        "Mở màn hình Register.",
                        "Nhập đầy đủ username, email, password và confirm password hợp lệ.",
                        "Chọn Account Type là Partner và nhấn Create Account.",
                    ],
                    [
                        "API nhận role partner và tạo tài khoản đối tác.",
                        "Màn hình chuyển về Login để người dùng đăng nhập.",
                        "Không tự gán role admin hoặc role ngoài customer/partner.",
                    ],
                ),
                (
                    "Không cho đăng ký email sai định dạng",
                    "Người dùng đang ở màn hình Register.",
                    [
                        "Nhập username hợp lệ.",
                        "Nhập email dạng sai, ví dụ abc hoặc user@.",
                        "Nhập password và confirm password hợp lệ rồi nhấn Create Account.",
                    ],
                    [
                        "Form hiển thị lỗi email không hợp lệ.",
                        "Không gọi API đăng ký hoặc API trả về lỗi validation 400.",
                        "Dữ liệu không được tạo trong bảng users.",
                    ],
                ),
                (
                    "Không cho đăng ký password dưới 6 ký tự",
                    "Người dùng đang ở màn hình Register.",
                    [
                        "Nhập username và email hợp lệ.",
                        "Nhập password có 5 ký tự.",
                        "Nhập confirm password giống password và nhấn Create Account.",
                    ],
                    [
                        "Form hiển thị thông báo password phải có ít nhất 6 ký tự.",
                        "Nút submit không tạo tài khoản thành công.",
                        "Người dùng vẫn ở màn hình Register.",
                    ],
                ),
                (
                    "Không cho đăng ký khi confirm password không khớp",
                    "Người dùng đang ở màn hình Register.",
                    [
                        "Nhập password hợp lệ.",
                        "Nhập confirm password khác password.",
                        "Nhấn Create Account.",
                    ],
                    [
                        "Form hiển thị lỗi Passwords do not match.",
                        "Không tạo tài khoản mới.",
                        "Các giá trị hợp lệ đã nhập vẫn được giữ để người dùng sửa.",
                    ],
                ),
                (
                    "Đăng nhập Customer thành công và điều hướng đúng",
                    "Tồn tại tài khoản customer đang active.",
                    [
                        "Mở màn hình Login.",
                        "Nhập email và password đúng.",
                        "Nhấn Login.",
                    ],
                    [
                        "API trả về access token, refresh token và thông tin user.",
                        "Ứng dụng lưu phiên đăng nhập an toàn.",
                        "Người dùng được điều hướng tới /customer/dashboard.",
                    ],
                ),
                (
                    "Đăng nhập Partner thành công và điều hướng đúng",
                    "Tồn tại tài khoản partner đang active.",
                    [
                        "Mở màn hình Login.",
                        "Nhập email và password đúng của partner.",
                        "Nhấn Login.",
                    ],
                    [
                        "Đăng nhập thành công.",
                        "Role partner được nhận diện đúng.",
                        "Người dùng được điều hướng tới /partner/dashboard.",
                    ],
                ),
                (
                    "Đăng nhập Admin thành công và điều hướng đúng",
                    "Tồn tại tài khoản admin, SUPER_ADMIN, OPERATOR hoặc ACCOUNTANT đang active.",
                    [
                        "Mở màn hình Login trên trình duyệt web.",
                        "Nhập email và password đúng của admin.",
                        "Nhấn Login.",
                    ],
                    [
                        "Đăng nhập thành công.",
                        "Ứng dụng điều hướng tới /admin/dashboard.",
                        "Menu admin hiển thị theo quyền của role đăng nhập.",
                    ],
                ),
                (
                    "Đăng nhập thất bại với sai mật khẩu",
                    "Tồn tại tài khoản active trong hệ thống.",
                    [
                        "Mở màn hình Login.",
                        "Nhập email đúng.",
                        "Nhập password sai và nhấn Login.",
                    ],
                    [
                        "API trả về lỗi xác thực phù hợp.",
                        "Màn hình hiển thị thông báo lỗi dễ hiểu.",
                        "Không cấp access token hoặc refresh token mới.",
                    ],
                ),
                (
                    "Ẩn và hiện mật khẩu trên màn hình Login",
                    "Người dùng đang ở màn hình Login.",
                    [
                        "Nhập giá trị vào ô Password.",
                        "Nhấn nút Show.",
                        "Nhấn lại nút Hide.",
                    ],
                    [
                        "Khi nhấn Show, mật khẩu được hiển thị dạng text.",
                        "Khi nhấn Hide, mật khẩu được che lại.",
                        "Giá trị password không bị mất trong quá trình chuyển trạng thái.",
                    ],
                ),
                (
                    "Refresh token hợp lệ cấp lại phiên",
                    "Người dùng đã đăng nhập và có refreshToken hợp lệ.",
                    [
                        "Gửi request refresh-token với refreshToken hợp lệ.",
                        "Quan sát response API.",
                        "Dùng access token mới gọi endpoint /me.",
                    ],
                    [
                        "API refresh-token trả về access token mới.",
                        "Endpoint /me nhận token mới và trả về đúng user hiện tại.",
                        "Không làm thay đổi role hoặc trạng thái người dùng.",
                    ],
                ),
                (
                    "Logout vô hiệu hóa phiên hiện tại",
                    "Người dùng đang đăng nhập.",
                    [
                        "Gọi chức năng Logout.",
                        "Sau khi logout, thử truy cập endpoint yêu cầu token.",
                        "Thử dùng refreshToken cũ để refresh.",
                    ],
                    [
                        "Ứng dụng xóa token khỏi storage.",
                        "Endpoint bảo vệ trả về 401 khi không có token hợp lệ.",
                        "RefreshToken cũ không tiếp tục cấp phiên mới nếu backend đã thu hồi.",
                    ],
                ),
            ],
        ),
        (
            "CUS_SEARCH",
            "Customer - Tìm kiếm & Khám phá",
            [
                (
                    "Hiển thị danh sách khách sạn đang active",
                    "Cơ sở dữ liệu có hotel_cards với isActive = true.",
                    [
                        "Mở tab Customer Dashboard hoặc trang danh sách khách sạn.",
                        "Gọi API GET /api/customer/hotels không truyền filter.",
                        "Quan sát danh sách trả về trên UI.",
                    ],
                    [
                        "Chỉ các khách sạn active được hiển thị.",
                        "Danh sách có tên, vị trí, ảnh, giá, rating và tag nếu có.",
                        "Không hiển thị khách sạn inactive.",
                    ],
                ),
                (
                    "Tìm kiếm theo tên khách sạn",
                    "Có ít nhất một khách sạn active có tên chứa từ khóa cần tìm.",
                    [
                        "Mở màn hình Search.",
                        "Nhập keyword trùng một phần tên khách sạn.",
                        "Thực hiện tìm kiếm.",
                    ],
                    [
                        "Danh sách kết quả chỉ gồm các khách sạn khớp từ khóa.",
                        "Tìm kiếm không phân biệt hoa thường.",
                        "Không phát sinh lỗi khi keyword có dấu tiếng Việt.",
                    ],
                ),
                (
                    "Tìm kiếm theo địa danh bỏ qua từ hành chính",
                    "Có dữ liệu khách sạn tại quận/huyện hoặc khu vực cụ thể.",
                    [
                        "Nhập keyword là tên địa danh có hoặc không có tiền tố Quận/Huyện.",
                        "Thực hiện tìm kiếm.",
                        "So sánh kết quả với dữ liệu khách sạn.",
                    ],
                    [
                        "Hệ thống tìm được khách sạn theo city, district, area hoặc location.",
                        "Các từ hành chính không làm sai lệch kết quả.",
                        "Kết quả không chứa khách sạn ngoài khu vực không liên quan.",
                    ],
                ),
                (
                    "Lọc khách sạn theo khoảng giá hợp lệ",
                    "Có dữ liệu hotel_cards với nhiều priceValue khác nhau.",
                    [
                        "Truyền minPrice và maxPrice hợp lệ vào bộ lọc.",
                        "Gọi API danh sách khách sạn.",
                        "Quan sát giá từng kết quả.",
                    ],
                    [
                        "Tất cả kết quả có priceValue nằm trong khoảng lọc.",
                        "Không trả về khách sạn có giá thấp hơn minPrice hoặc cao hơn maxPrice.",
                        "Tổng số kết quả phản ánh đúng dữ liệu sau lọc.",
                    ],
                ),
                (
                    "Từ chối khoảng giá có maxPrice nhỏ hơn minPrice",
                    "Người dùng đang ở màn hình tìm kiếm hoặc gọi API trực tiếp.",
                    [
                        "Nhập minPrice lớn hơn maxPrice.",
                        "Thực hiện tìm kiếm.",
                        "Quan sát response hoặc lỗi form.",
                    ],
                    [
                        "Hệ thống trả lỗi validation rõ ràng.",
                        "Không trả về danh sách sai.",
                        "UI giữ lại dữ liệu để người dùng sửa bộ lọc.",
                    ],
                ),
                (
                    "Sắp xếp theo rating giảm dần",
                    "Có nhiều khách sạn active với rating khác nhau.",
                    [
                        "Chọn sort = rating.",
                        "Gọi API hoặc thao tác sắp xếp trên UI.",
                        "Kiểm tra thứ tự danh sách.",
                    ],
                    [
                        "Khách sạn có rating cao hơn đứng trước.",
                        "Các thông tin card không bị mất sau khi sắp xếp.",
                        "Không thay đổi bộ lọc đang áp dụng.",
                    ],
                ),
                (
                    "Sắp xếp giá tăng dần",
                    "Có nhiều khách sạn active với priceValue khác nhau.",
                    [
                        "Chọn sort = price-asc.",
                        "Thực hiện tìm kiếm.",
                        "Kiểm tra thứ tự giá.",
                    ],
                    [
                        "Danh sách được sắp xếp từ giá thấp đến cao.",
                        "Các khách sạn không active vẫn bị loại bỏ.",
                        "Không hiển thị dữ liệu trùng lặp.",
                    ],
                ),
                (
                    "Sắp xếp giá giảm dần",
                    "Có nhiều khách sạn active với priceValue khác nhau.",
                    [
                        "Chọn sort = price-desc.",
                        "Thực hiện tìm kiếm.",
                        "Kiểm tra thứ tự giá.",
                    ],
                    [
                        "Danh sách được sắp xếp từ giá cao đến thấp.",
                        "Giá hiển thị trên UI khớp với priceValue backend.",
                        "Không lỗi khi kết quả ít hơn một trang.",
                    ],
                ),
                (
                    "Giới hạn limit tối đa 50 kết quả",
                    "Có hơn 50 khách sạn active.",
                    [
                        "Gọi API với limit lớn hơn 50.",
                        "Quan sát response.",
                        "Đếm số phần tử trả về.",
                    ],
                    [
                        "Validation không cho limit vượt quá 50 hoặc backend giới hạn tối đa 50.",
                        "Response không trả quá 50 khách sạn.",
                        "Hiệu năng vẫn ổn định.",
                    ],
                ),
                (
                    "Lọc khách sạn theo tiện nghi phòng",
                    "Có dữ liệu room amenities gắn với một số khách sạn.",
                    [
                        "Chọn một hoặc nhiều tiện nghi phòng.",
                        "Gọi API với roomAmenities.",
                        "Quan sát danh sách kết quả.",
                    ],
                    [
                        "Chỉ khách sạn có loại phòng chứa tiện nghi đã chọn được hiển thị.",
                        "Thông tin tiện nghi được attach vào card khi có dữ liệu.",
                        "Không lỗi khi tiện nghi không có khách sạn nào phù hợp.",
                    ],
                ),
                (
                    "Xem danh sách địa điểm khách sạn",
                    "Có dữ liệu hotel_addresses hoặc hotel_cards.",
                    [
                        "Gọi GET /api/customer/hotels/locations.",
                        "Mở modal/chức năng chọn địa điểm trên UI.",
                        "Kiểm tra cây tỉnh/thành, quận/huyện, phường/xã.",
                    ],
                    [
                        "API trả về cây địa điểm hợp lệ.",
                        "Nếu hotel_addresses trống, hệ thống fallback từ hotel_cards.",
                        "UI không hiển thị địa điểm rỗng hoặc trùng không cần thiết.",
                    ],
                ),
                (
                    "Hiển thị trạng thái rỗng khi không có kết quả",
                    "Không có khách sạn nào khớp bộ lọc đã chọn.",
                    [
                        "Nhập keyword hoặc filter không khớp dữ liệu.",
                        "Thực hiện tìm kiếm.",
                        "Quan sát UI.",
                    ],
                    [
                        "UI hiển thị trạng thái không có kết quả rõ ràng.",
                        "Không crash hoặc hiển thị dữ liệu cũ gây hiểu nhầm.",
                        "Người dùng có thể sửa hoặc xóa filter để tìm lại.",
                    ],
                ),
            ],
        ),
        (
            "HOTEL",
            "Customer - Chi tiết & Phòng trống",
            [
                (
                    "Xem chi tiết khách sạn hợp lệ",
                    "Tồn tại khách sạn active với id hợp lệ.",
                    [
                        "Từ danh sách khách sạn, chọn một khách sạn.",
                        "Gọi GET /api/customer/hotels/:id.",
                        "Quan sát màn hình chi tiết.",
                    ],
                    [
                        "Màn hình hiển thị tên, ảnh, địa chỉ, rating, tag và thông tin giá.",
                        "API trả đúng khách sạn theo id.",
                        "Không hiển thị khách sạn inactive.",
                    ],
                ),
                (
                    "Trả lỗi khi xem khách sạn không tồn tại",
                    "Không có khách sạn active với id được dùng.",
                    [
                        "Gọi GET /api/customer/hotels/:id với id không tồn tại.",
                        "Quan sát response API.",
                        "Quan sát trạng thái UI nếu truy cập bằng link.",
                    ],
                    [
                        "API trả lỗi 404 HOTEL_NOT_FOUND.",
                        "UI hiển thị thông báo không tìm thấy khách sạn.",
                        "Không chuyển người dùng tới màn hình đặt phòng sai.",
                    ],
                ),
                (
                    "Ghi nhận khách sạn đã xem",
                    "Customer đã đăng nhập và khách sạn active tồn tại.",
                    [
                        "Gọi POST /api/customer/hotels/:id/view.",
                        "Gọi GET /api/customer/hotels/viewed.",
                        "Kiểm tra thứ tự danh sách đã xem.",
                    ],
                    [
                        "Khách sạn được thêm hoặc cập nhật viewedAt.",
                        "Danh sách đã xem hiển thị khách sạn mới nhất trước.",
                        "Không tạo bản ghi trùng cho cùng user và hotel.",
                    ],
                ),
                (
                    "Không cho ghi nhận viewed hotel khi chưa đăng nhập",
                    "Người dùng chưa có token customer.",
                    [
                        "Gọi POST /api/customer/hotels/:id/view không gửi token.",
                        "Quan sát response.",
                        "Kiểm tra dữ liệu customer_viewed_hotels.",
                    ],
                    [
                        "API trả 401 Unauthorized.",
                        "Không ghi dữ liệu viewed hotel.",
                        "UI yêu cầu đăng nhập nếu cần lưu lịch sử.",
                    ],
                ),
                (
                    "Lấy danh sách phòng của khách sạn",
                    "Khách sạn có room types active.",
                    [
                        "Mở chi tiết khách sạn.",
                        "Gọi GET /api/customer/hotels/:id/rooms.",
                        "Quan sát danh sách phòng.",
                    ],
                    [
                        "Các loại phòng active được hiển thị.",
                        "Thông tin sức chứa, giường, ảnh, tiện nghi và giá được hiển thị đúng.",
                        "Phòng inactive không được ưu tiên hiển thị cho khách đặt.",
                    ],
                ),
                (
                    "Kiểm tra slot theo giờ trong ngày hiện tại",
                    "Khách sạn có phòng active và ngày kiểm tra là hôm nay.",
                    [
                        "Gọi API availability với bookingType = Theo giờ và date là hôm nay.",
                        "Quan sát các slot trước thời điểm hiện tại.",
                        "Quan sát các slot sau thời điểm hiện tại.",
                    ],
                    [
                        "Các slot đã qua không khả dụng.",
                        "Slot tương lai được tính theo mốc 00 hoặc 30 phút tiếp theo.",
                        "Mỗi slot có available và maxHours hợp lệ.",
                    ],
                ),
                (
                    "Kiểm tra slot theo giờ ngày tương lai",
                    "Khách sạn có phòng active và ngày kiểm tra là ngày tương lai.",
                    [
                        "Gọi availability với bookingType = Theo giờ và date tương lai.",
                        "Quan sát toàn bộ slot 00:00 đến 23:30.",
                        "Kiểm tra maxHours tại các khung cuối ngày.",
                    ],
                    [
                        "API trả các slot cách nhau 30 phút.",
                        "maxHours không vượt quá 10 giờ và không vượt quá cuối ngày.",
                        "Slot có phòng bị đặt hết sẽ trả available = false.",
                    ],
                ),
                (
                    "Kiểm tra slot Qua đêm",
                    "Khách sạn có ít nhất một room type active.",
                    [
                        "Gọi availability với bookingType = Qua đêm.",
                        "Quan sát response.",
                        "So sánh khoảng check-in/check-out backend sử dụng.",
                    ],
                    [
                        "API trả một slot 22:00.",
                        "Khoảng đặt phòng được tính 12 giờ từ 22:00.",
                        "available phản ánh đúng số phòng còn trống.",
                    ],
                ),
                (
                    "Kiểm tra slot Theo ngày",
                    "Khách sạn có ít nhất một room type active.",
                    [
                        "Gọi availability với bookingType = Theo ngày.",
                        "Quan sát response.",
                        "Kiểm tra thời gian slot.",
                    ],
                    [
                        "API trả một slot 14:00.",
                        "Khoảng đặt phòng được tính từ 14:00 và phù hợp logic backend.",
                        "available phản ánh đúng inventory/booking hiện có.",
                    ],
                ),
                (
                    "Từ chối date sai định dạng khi kiểm tra availability",
                    "Người dùng hoặc client gọi API availability.",
                    [
                        "Gọi availability với date không theo YYYY-MM-DD.",
                        "Quan sát response API.",
                        "Kiểm tra UI khi nhập ngày sai nếu có.",
                    ],
                    [
                        "API trả lỗi validation.",
                        "Không thực hiện truy vấn phòng trống với ngày sai.",
                        "UI hiển thị lỗi rõ ràng hoặc không cho chọn ngày sai.",
                    ],
                ),
                (
                    "Gallery ảnh fallback khi phòng/khách sạn thiếu ảnh",
                    "Có khách sạn hoặc phòng thiếu ảnh media.",
                    [
                        "Mở chi tiết khách sạn/phòng thiếu ảnh.",
                        "Quan sát ảnh chính và gallery.",
                        "Thử mở ảnh trong viewer nếu có.",
                    ],
                    [
                        "UI hiển thị ảnh fallback thay vì vỡ layout.",
                        "Không có icon ảnh lỗi hoặc vùng trắng bất thường.",
                        "Viewer không crash khi danh sách ảnh rỗng.",
                    ],
                ),
                (
                    "Responsive màn hình chi tiết khách sạn web và mobile",
                    "Ứng dụng chạy trên trình duyệt web và thiết bị mobile/emulator.",
                    [
                        "Mở chi tiết cùng một khách sạn trên web desktop.",
                        "Mở trên viewport mobile.",
                        "Cuộn qua gallery, tiện nghi và danh sách phòng.",
                    ],
                    [
                        "Nội dung không bị chồng lấn hoặc tràn ngoài màn hình.",
                        "Các nút đặt phòng vẫn dễ thao tác.",
                        "Ảnh, text giá và thông tin phòng giữ thứ tự đọc hợp lý.",
                    ],
                ),
            ],
        ),
        (
            "BOOKING",
            "Customer - Đặt phòng",
            [
                (
                    "Tạo booking Theo giờ hợp lệ với VietQR",
                    "Customer đã đăng nhập, khách sạn và room type còn phòng.",
                    [
                        "Chọn khách sạn và loại phòng còn trống.",
                        "Chọn bookingType Theo giờ, checkIn/checkOut hợp lệ, guests trong giới hạn.",
                        "Chọn VIETQR và xác nhận đặt phòng.",
                    ],
                    [
                        "API tạo booking thành công với status phù hợp.",
                        "bookingCode được sinh theo định dạng BKyyyyMMddxxxx.",
                        "Payment VietQR được tạo với paymentCode, content và QR URL.",
                    ],
                ),
                (
                    "Tạo booking Qua đêm hợp lệ",
                    "Customer đã đăng nhập và có phòng trống qua đêm.",
                    [
                        "Chọn bookingType Qua đêm.",
                        "Chọn ngày và thông tin khách hợp lệ.",
                        "Xác nhận đặt phòng.",
                    ],
                    [
                        "Booking được lưu với bookingType overnight trong database.",
                        "Thời gian nhận/trả phòng hợp lệ và checkOut sau checkIn.",
                        "UI chuyển sang bước thanh toán hoặc xác nhận theo payment method.",
                    ],
                ),
                (
                    "Tạo booking Theo ngày hợp lệ",
                    "Customer đã đăng nhập và có phòng trống theo ngày.",
                    [
                        "Chọn bookingType Theo ngày.",
                        "Nhập số khách và thông tin liên hệ hợp lệ.",
                        "Xác nhận booking.",
                    ],
                    [
                        "Booking được tạo với bookingType daily.",
                        "Tổng tiền là số nguyên dương.",
                        "Thông tin khách hàng được lưu nếu có nhập.",
                    ],
                ),
                (
                    "Tạo booking với Pay at hotel",
                    "Customer đã đăng nhập, khách sạn cho phép đặt phòng và trả tại khách sạn.",
                    [
                        "Chọn phòng và thời gian hợp lệ.",
                        "Chọn paymentMethod = PAY_AT_HOTEL.",
                        "Xác nhận booking.",
                    ],
                    [
                        "Booking được tạo thành công.",
                        "Payment method là PAY_AT_HOTEL và không có QR URL/content ngân hàng.",
                        "Trạng thái hiển thị là chờ nhận phòng hoặc trạng thái tương đương.",
                    ],
                ),
                (
                    "Từ chối booking khi checkOut không sau checkIn",
                    "Customer đã đăng nhập.",
                    [
                        "Chọn hoặc gửi checkOut bằng hoặc trước checkIn.",
                        "Nhấn xác nhận booking.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả lỗi validation checkOut phải sau checkIn.",
                        "Không tạo booking hoặc payment.",
                        "UI hiển thị lỗi để người dùng sửa thời gian.",
                    ],
                ),
                (
                    "Từ chối booking khi guests nhỏ hơn 1",
                    "Customer đã đăng nhập.",
                    [
                        "Gửi request tạo booking với guests = 0.",
                        "Quan sát response API.",
                        "Kiểm tra dữ liệu booking.",
                    ],
                    [
                        "API trả lỗi validation.",
                        "Không tạo booking.",
                        "UI không cho giảm số khách dưới 1.",
                    ],
                ),
                (
                    "Từ chối booking khi guests vượt quá 20",
                    "Customer đã đăng nhập.",
                    [
                        "Gửi request tạo booking với guests = 21.",
                        "Quan sát response.",
                        "Kiểm tra UI input số khách.",
                    ],
                    [
                        "API trả lỗi validation max 20.",
                        "Không tạo booking.",
                        "UI giới hạn hoặc cảnh báo khi vượt giới hạn.",
                    ],
                ),
                (
                    "Từ chối booking khi amount không dương",
                    "Customer đã đăng nhập.",
                    [
                        "Gửi request tạo booking với amount = 0 hoặc âm.",
                        "Quan sát response API.",
                        "Kiểm tra bảng Booking và Payment.",
                    ],
                    [
                        "API trả lỗi tổng tiền không hợp lệ.",
                        "Không tạo booking.",
                        "Không tạo payment rác.",
                    ],
                ),
                (
                    "Từ chối booking khi phòng đã hết",
                    "Room type đã được đặt hết trong khoảng thời gian chọn.",
                    [
                        "Chọn khung giờ/ngày đã hết phòng.",
                        "Thực hiện đặt phòng.",
                        "Quan sát response và UI.",
                    ],
                    [
                        "Backend kiểm tra số phòng đã giữ trước khi tạo booking.",
                        "Booking bị từ chối với thông báo hết phòng hoặc không còn phòng trống.",
                        "Không overbooking vượt totalUnits.",
                    ],
                ),
                (
                    "Xem danh sách booking của tôi",
                    "Customer đã có nhiều booking ở các trạng thái khác nhau.",
                    [
                        "Đăng nhập customer.",
                        "Mở màn hình Bookings.",
                        "Gọi GET /api/customer/bookings.",
                    ],
                    [
                        "Chỉ booking của user hiện tại được hiển thị.",
                        "Thông tin mã booking, khách sạn, phòng, thời gian, giá và trạng thái đúng.",
                        "Booking mới nhất được sắp xếp hợp lý.",
                    ],
                ),
                (
                    "Xem chi tiết booking thuộc tài khoản hiện tại",
                    "Customer đã có ít nhất một booking.",
                    [
                        "Mở danh sách Bookings.",
                        "Chọn một booking.",
                        "Quan sát màn hình chi tiết.",
                    ],
                    [
                        "Chi tiết trả về đúng booking của user hiện tại.",
                        "Có thông tin phòng, khách sạn, thanh toán và hỗ trợ nếu cần.",
                        "Không hiển thị dữ liệu booking của user khác.",
                    ],
                ),
                (
                    "Không cho xem booking của user khác",
                    "Có booking thuộc customer khác.",
                    [
                        "Đăng nhập bằng customer A.",
                        "Gọi GET /api/customer/bookings/:id với id booking của customer B.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả 404 hoặc 403 theo thiết kế bảo mật.",
                        "Không rò rỉ thông tin khách hàng, số điện thoại hoặc thanh toán của user khác.",
                        "UI hiển thị thông báo không tìm thấy hoặc không có quyền.",
                    ],
                ),
                (
                    "Hủy booking đang chờ xử lý",
                    "Customer có booking ở trạng thái có thể hủy.",
                    [
                        "Mở chi tiết booking.",
                        "Nhấn Cancel booking.",
                        "Xác nhận hủy.",
                    ],
                    [
                        "Booking chuyển sang CANCELLED.",
                        "Nếu có voucher đã dùng, usedCount được hoàn lại theo logic backend.",
                        "Danh sách booking cập nhật trạng thái đã hủy.",
                    ],
                ),
            ],
        ),
        (
            "PAY",
            "Thanh toán VietQR & Sepay",
            [
                (
                    "Hiển thị QR thanh toán mới tạo",
                    "Booking VietQR vừa được tạo thành công.",
                    [
                        "Đi tới màn hình BookingPayment.",
                        "Quan sát thông tin QR, số tiền, nội dung chuyển khoản và ngân hàng.",
                        "So sánh với response payment.",
                    ],
                    [
                        "QR URL được hiển thị đúng.",
                        "Số tiền, paymentCode, bookingCode và nội dung chuyển khoản khớp backend.",
                        "Countdown bắt đầu từ 15 phút hoặc từ expiresAt backend.",
                    ],
                ),
                (
                    "Countdown chuyển sang giai đoạn grace sau 15 phút",
                    "Có payment VietQR ở trạng thái PENDING.",
                    [
                        "Mở màn hình thanh toán.",
                        "Giả lập thời gian vượt expiresAt nhưng chưa vượt graceExpiresAt.",
                        "Gọi kiểm tra payment status.",
                    ],
                    [
                        "Payment phase chuyển sang GRACE.",
                        "UI thông báo hệ thống tiếp tục kiểm tra giao dịch tối đa 5 phút.",
                        "Chưa cho kết luận paid nếu chưa có webhook hợp lệ.",
                    ],
                ),
                (
                    "Payment hết hạn sau grace period",
                    "Có payment VietQR PENDING đã quá graceExpiresAt.",
                    [
                        "Gọi endpoint payment-status sau khi hết grace.",
                        "Quan sát status trả về.",
                        "Quan sát UI thanh toán.",
                    ],
                    [
                        "Payment được final thành EXPIRED_FINAL hoặc trạng thái tương đương.",
                        "UI hiển thị không ghi nhận thanh toán.",
                        "Người dùng được cung cấp lựa chọn tạo QR mới hoặc liên hệ hỗ trợ nếu có.",
                    ],
                ),
                (
                    "Tạo QR mới cho booking chưa thanh toán",
                    "Booking VietQR có payment đã hết hạn và backend cho phép tạo QR mới.",
                    [
                        "Nhấn nút tạo QR mới.",
                        "Gọi POST /api/customer/bookings/:id/payment/new-qr.",
                        "Quan sát session thanh toán mới.",
                    ],
                    [
                        "Backend tạo payment attempt mới với attemptNo tăng lên.",
                        "QR mới có paymentCode và expiresAt mới.",
                        "UI reset countdown và không dùng lại QR cũ.",
                    ],
                ),
                (
                    "Webhook Sepay hợp lệ ghi nhận thanh toán",
                    "Tồn tại payment PENDING với bookingCode/paymentCode đúng.",
                    [
                        "Gửi payload Sepay có transferType nhận tiền, accountNumber đúng, amount đủ và nội dung chứa mã đúng.",
                        "Gọi POST /api/customer/payments/sepay/webhook.",
                        "Kiểm tra payment và booking sau xử lý.",
                    ],
                    [
                        "Webhook được log processed = true.",
                        "Payment chuyển PAID và lưu paidAt/referenceCode nếu có.",
                        "Booking chuyển sang trạng thái xác nhận phù hợp.",
                    ],
                ),
                (
                    "Webhook sai số tiền không ghi nhận thanh toán",
                    "Tồn tại payment PENDING với amount yêu cầu lớn hơn transferAmount.",
                    [
                        "Gửi payload Sepay có mã đúng nhưng transferAmount thấp hơn amount.",
                        "Quan sát response và payment.",
                        "Kiểm tra webhook log.",
                    ],
                    [
                        "Payment không chuyển PAID.",
                        "failureReason là INVALID_AMOUNT hoặc thông báo tương đương.",
                        "Webhook log ghi nhận processed = false để hỗ trợ đối soát.",
                    ],
                ),
                (
                    "Webhook sai nội dung chuyển khoản",
                    "Tồn tại booking/payment PENDING.",
                    [
                        "Gửi payload Sepay không chứa bookingCode/paymentCode hợp lệ.",
                        "Quan sát xử lý backend.",
                        "Kiểm tra log webhook.",
                    ],
                    [
                        "Backend không match payment hợp lệ.",
                        "Không cập nhật payment của booking khác.",
                        "Webhook log ghi lỗi PAYMENT_NOT_FOUND hoặc INVALID_CONTENT phù hợp.",
                    ],
                ),
                (
                    "Webhook gửi vào sai tài khoản nhận",
                    "Tồn tại payment PENDING và cấu hình tài khoản nhận của StayHub.",
                    [
                        "Gửi payload Sepay có accountNumber khác tài khoản nhận.",
                        "Quan sát response.",
                        "Kiểm tra payment.",
                    ],
                    [
                        "Payment không chuyển PAID.",
                        "failureReason thể hiện INVALID_ACCOUNT hoặc lỗi tương đương.",
                        "Dữ liệu giao dịch được lưu trong webhook log để đối soát.",
                    ],
                ),
                (
                    "Không ghi nhận webhook đến sau khi payment đã hết hạn cuối",
                    "Payment đã ở trạng thái EXPIRED_FINAL hoặc PAYMENT_NOT_RECORDED.",
                    [
                        "Gửi payload Sepay sau grace period.",
                        "Quan sát payment status.",
                        "Kiểm tra thông tin failure.",
                    ],
                    [
                        "Payment không quay lại PAID tự động nếu đã final theo policy.",
                        "Backend ghi nhận giao dịch cần hỗ trợ thủ công.",
                        "UI hiển thị hướng dẫn liên hệ support.",
                    ],
                ),
                (
                    "Kiểm tra trạng thái payment đã PAID trên UI",
                    "Payment đã được webhook xác nhận PAID.",
                    [
                        "Ở màn hình thanh toán, nhấn kiểm tra trạng thái.",
                        "Gọi GET /api/customer/bookings/:id/payment-status.",
                        "Quan sát UI.",
                    ],
                    [
                        "API trả isPaid = true.",
                        "UI hiển thị trạng thái xác nhận thành công.",
                        "Người dùng có thể chuyển tới chi tiết booking.",
                    ],
                ),
                (
                    "Không hiển thị QR cho Pay at hotel",
                    "Booking được tạo với paymentMethod PAY_AT_HOTEL.",
                    [
                        "Tạo booking trả tại khách sạn.",
                        "Mở chi tiết booking hoặc màn hình thanh toán.",
                        "Quan sát các trường payment.",
                    ],
                    [
                        "Không hiển thị QR URL, bankCode, accountNumber hoặc content.",
                        "Payment method hiển thị Pay at hotel.",
                        "Không có countdown hết hạn thanh toán online.",
                    ],
                ),
                (
                    "Màn hình thanh toán xử lý session không hợp lệ",
                    "Người dùng mở màn hình payment bằng URL thiếu hoặc sai param session.",
                    [
                        "Truy cập màn hình BookingPayment với session rỗng hoặc JSON sai.",
                        "Quan sát UI.",
                        "Thử quay lại màn hình trước.",
                    ],
                    [
                        "Ứng dụng không crash khi parse session lỗi.",
                        "UI hiển thị lỗi hoặc trạng thái không tìm thấy phiên thanh toán.",
                        "Người dùng có thể quay lại an toàn.",
                    ],
                ),
            ],
        ),
        (
            "VOUCHER",
            "Voucher & Ưu đãi",
            [
                (
                    "Liệt kê voucher khả dụng cho khách sạn",
                    "Khách sạn có voucher active của hệ thống và voucher riêng khách sạn.",
                    [
                        "Customer đăng nhập.",
                        "Gọi GET /api/customer/hotels/:id/vouchers.",
                        "Quan sát danh sách voucher.",
                    ],
                    [
                        "API trả voucher active thuộc hotel hoặc voucher hệ thống hotelId null.",
                        "Voucher riêng của khách sạn được ưu tiên khi trùng code.",
                        "Voucher không thỏa điều kiện bị loại khỏi danh sách khả dụng.",
                    ],
                ),
                (
                    "Áp dụng voucher giảm phần trăm hợp lệ",
                    "Có voucher active action percent và đơn hàng đủ điều kiện.",
                    [
                        "Nhập mã voucher percent trên màn hình đặt phòng.",
                        "Gọi validate voucher với subtotal hợp lệ.",
                        "Quan sát tổng tiền sau giảm.",
                    ],
                    [
                        "discount = subtotal * percent / 100.",
                        "finalTotal = subtotal - discount.",
                        "Discount không vượt quá max nếu action có max.",
                    ],
                ),
                (
                    "Áp dụng voucher giảm tiền cố định hợp lệ",
                    "Có voucher active action fixed.",
                    [
                        "Nhập mã voucher fixed.",
                        "Gọi validate voucher.",
                        "Quan sát kết quả tính tiền.",
                    ],
                    [
                        "discount bằng giá trị fixed.",
                        "finalTotal không âm nếu fixed lớn hơn subtotal.",
                        "Thông tin discountType/discountValue legacy vẫn đúng cho frontend.",
                    ],
                ),
                (
                    "Từ chối voucher khi chưa đạt minOrder",
                    "Có voucher rule minOrder lớn hơn subtotal hiện tại.",
                    [
                        "Nhập voucher có rule minOrder.",
                        "Subtotal nhỏ hơn giá trị minOrder.",
                        "Gọi validate voucher.",
                    ],
                    [
                        "API trả lỗi đơn hàng chưa đạt giá trị tối thiểu.",
                        "Không áp dụng discount vào booking.",
                        "UI hiển thị lý do để người dùng hiểu.",
                    ],
                ),
                (
                    "Từ chối voucher hết hạn",
                    "Có voucher active nhưng constraints.endDate nhỏ hơn thời điểm hiện tại.",
                    [
                        "Nhập mã voucher đã hết hạn.",
                        "Gọi validate voucher.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả lỗi voucher đã hết hạn.",
                        "Voucher không xuất hiện trong danh sách khả dụng.",
                        "Booking không được giảm giá bởi voucher này.",
                    ],
                ),
                (
                    "Từ chối voucher chưa đến ngày bắt đầu",
                    "Có voucher có constraints.startDate trong tương lai.",
                    [
                        "Nhập mã voucher chưa bắt đầu.",
                        "Gọi validate voucher.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả lỗi voucher chưa đến thời gian sử dụng.",
                        "Không áp dụng discount.",
                        "UI không hiển thị như voucher có thể dùng ngay.",
                    ],
                ),
                (
                    "Từ chối voucher hết số lượt sử dụng",
                    "Voucher có usedCount >= usageLimit.",
                    [
                        "Nhập mã voucher hết lượt.",
                        "Gọi validate voucher.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả lỗi voucher đã hết lượt sử dụng.",
                        "Không tăng usedCount thêm.",
                        "UI hiển thị trạng thái hết lượt.",
                    ],
                ),
                (
                    "Từ chối voucher vượt giới hạn perUser",
                    "Customer đã dùng voucher đạt constraints.perUser.",
                    [
                        "Đăng nhập customer đã dùng voucher trước đó.",
                        "Nhập lại cùng mã voucher.",
                        "Gọi validate voucher.",
                    ],
                    [
                        "API trả lỗi người dùng đã sử dụng quá số lần cho phép.",
                        "Không áp dụng discount.",
                        "Không tạo booking giảm giá bằng voucher đó.",
                    ],
                ),
                (
                    "Từ chối voucher sai loại booking",
                    "Voucher có rule bookingType khác loại booking đang chọn.",
                    [
                        "Chọn bookingType Theo ngày.",
                        "Nhập voucher chỉ áp dụng hourly.",
                        "Gọi validate voucher.",
                    ],
                    [
                        "API trả lỗi voucher không áp dụng cho kiểu đặt phòng này.",
                        "Tổng tiền giữ nguyên.",
                        "UI cho phép chọn voucher khác.",
                    ],
                ),
                (
                    "Từ chối voucher sai loại phòng",
                    "Voucher có rule roomType không chứa roomTypeId đang đặt.",
                    [
                        "Chọn một room type không nằm trong danh sách áp dụng.",
                        "Nhập mã voucher.",
                        "Gọi validate voucher.",
                    ],
                    [
                        "API trả lỗi voucher không áp dụng cho loại phòng này.",
                        "Không áp dụng giảm giá.",
                        "Không tăng usedCount.",
                    ],
                ),
                (
                    "Voucher firstBooking chỉ áp dụng lần đặt đầu tiên",
                    "Có voucher rule firstBooking.",
                    [
                        "Customer chưa từng booking áp dụng voucher.",
                        "Validate voucher và tạo booking thành công.",
                        "Customer có booking trước đó thử validate lại voucher.",
                    ],
                    [
                        "Lần đầu được áp dụng nếu các điều kiện khác hợp lệ.",
                        "Lần sau bị từ chối với lý do chỉ áp dụng lần đặt đầu tiên.",
                        "Kết quả dựa trên hasPreviousBooking chính xác.",
                    ],
                ),
                (
                    "Hoàn lại lượt voucher khi hủy booking",
                    "Booking có voucherCode và voucher đã tăng usedCount.",
                    [
                        "Tạo booking thành công với voucher.",
                        "Kiểm tra usedCount tăng.",
                        "Hủy booking.",
                    ],
                    [
                        "usedCount giảm đi 1 nhưng không nhỏ hơn 0.",
                        "Booking chuyển CANCELLED.",
                        "Voucher có thể tiếp tục dùng nếu vẫn còn lượt.",
                    ],
                ),
            ],
        ),
        (
            "CUS_ACCOUNT",
            "Customer - Tài khoản, Tin nhắn & Hỗ trợ",
            [
                (
                    "Xem hồ sơ cá nhân",
                    "Customer đã đăng nhập.",
                    [
                        "Mở màn hình Profile.",
                        "Gọi GET /api/customer/profile/me.",
                        "Quan sát dữ liệu hiển thị.",
                    ],
                    [
                        "Hiển thị đúng username, email, phone/avatar nếu có.",
                        "Không hiển thị password hoặc refreshToken.",
                        "Nếu token hết hạn, UI yêu cầu đăng nhập lại.",
                    ],
                ),
                (
                    "Cập nhật hồ sơ cá nhân hợp lệ",
                    "Customer đã đăng nhập.",
                    [
                        "Mở chỉnh sửa hồ sơ.",
                        "Cập nhật username hoặc phone hợp lệ.",
                        "Lưu thay đổi.",
                    ],
                    [
                        "API PATCH profile/me cập nhật thành công.",
                        "UI hiển thị dữ liệu mới sau khi lưu.",
                        "Không thay đổi email/role nếu không được phép.",
                    ],
                ),
                (
                    "Không cho cập nhật phone trùng",
                    "Đã tồn tại user khác có phone cần cập nhật.",
                    [
                        "Customer nhập phone đã được dùng bởi tài khoản khác.",
                        "Nhấn lưu.",
                        "Quan sát response.",
                    ],
                    [
                        "Backend trả lỗi unique hoặc thông báo phù hợp.",
                        "Hồ sơ hiện tại không bị cập nhật sai.",
                        "UI giữ form để người dùng sửa.",
                    ],
                ),
                (
                    "Xem danh sách tin nhắn",
                    "Customer đã đăng nhập và có customer_messages.",
                    [
                        "Mở màn hình Messages.",
                        "Gọi GET /api/customer/messages.",
                        "Quan sát danh sách.",
                    ],
                    [
                        "Tin nhắn của user hiện tại được hiển thị.",
                        "Có trạng thái đã đọc/chưa đọc.",
                        "Sắp xếp theo sortOrder hoặc thời gian phù hợp.",
                    ],
                ),
                (
                    "Đánh dấu tin nhắn đã đọc",
                    "Customer có tin nhắn chưa đọc.",
                    [
                        "Mở danh sách Messages.",
                        "Chọn một tin nhắn chưa đọc.",
                        "Gọi PATCH /api/customer/messages/:id/read.",
                    ],
                    [
                        "isRead chuyển true.",
                        "Badge hoặc số tin chưa đọc giảm tương ứng.",
                        "Không ảnh hưởng tin nhắn của user khác.",
                    ],
                ),
                (
                    "Xem danh sách thông báo",
                    "Customer có customer_notifications.",
                    [
                        "Mở màn hình Notifications.",
                        "Gọi GET /api/customer/notifications.",
                        "Quan sát tab/type nếu có.",
                    ],
                    [
                        "Thông báo booking/offers/others hiển thị đúng.",
                        "Thông báo chưa đọc được phân biệt rõ.",
                        "Không hiển thị thông báo của user khác.",
                    ],
                ),
                (
                    "Đánh dấu tất cả thông báo đã đọc",
                    "Customer có nhiều thông báo chưa đọc.",
                    [
                        "Mở Notifications.",
                        "Chọn Mark all as read.",
                        "Gọi PATCH /api/customer/notifications/read-all.",
                    ],
                    [
                        "Tất cả thông báo của customer hiện tại chuyển isRead true.",
                        "Badge chưa đọc về 0.",
                        "Không xóa thông báo khỏi danh sách.",
                    ],
                ),
                (
                    "Xóa một thông báo",
                    "Customer có ít nhất một thông báo.",
                    [
                        "Mở Notifications.",
                        "Chọn xóa một thông báo.",
                        "Xác nhận thao tác.",
                    ],
                    [
                        "Thông báo được xóa khỏi danh sách của user hiện tại.",
                        "API trả thành công.",
                        "Không xóa thông báo khác ngoài id đã chọn.",
                    ],
                ),
                (
                    "Xóa tất cả thông báo",
                    "Customer có nhiều thông báo.",
                    [
                        "Mở Notifications.",
                        "Chọn xóa tất cả.",
                        "Xác nhận.",
                    ],
                    [
                        "Danh sách thông báo của user hiện tại rỗng.",
                        "UI hiển thị trạng thái empty state.",
                        "Không ảnh hưởng dữ liệu message.",
                    ],
                ),
                (
                    "Mở FAQ và điều khoản hỗ trợ",
                    "Ứng dụng customer đã cài route support/faqs và support/terms.",
                    [
                        "Mở màn hình Support.",
                        "Chọn FAQs.",
                        "Chọn Terms.",
                    ],
                    [
                        "Các màn hình hỗ trợ mở đúng route.",
                        "Nội dung hiển thị dễ đọc trên mobile và web.",
                        "Nút quay lại hoạt động đúng.",
                    ],
                ),
                (
                    "Gửi liên hệ hỗ trợ với dữ liệu hợp lệ",
                    "Customer mở màn hình Contact Support.",
                    [
                        "Nhập chủ đề, nội dung và thông tin liên hệ hợp lệ.",
                        "Nhấn gửi.",
                        "Quan sát phản hồi UI.",
                    ],
                    [
                        "Form không báo lỗi validation.",
                        "Người dùng nhận được thông báo gửi thành công hoặc hướng dẫn liên hệ.",
                        "Nội dung dài được wrap, không làm vỡ layout.",
                    ],
                ),
                (
                    "Chuyển theme sáng/tối trên Customer UI",
                    "Ứng dụng đang ở khu vực customer có ThemeToggle.",
                    [
                        "Nhấn nút đổi theme.",
                        "Quan sát màu nền, chữ và card.",
                        "Điều hướng sang màn hình khác.",
                    ],
                    [
                        "Theme đổi nhất quán toàn khu vực customer.",
                        "Độ tương phản chữ/nền vẫn đọc được.",
                        "Trạng thái theme được giữ theo context/storage nếu có hỗ trợ.",
                    ],
                ),
            ],
        ),
        (
            "PARTNER_HOTEL",
            "Partner - Quản lý khách sạn",
            [
                (
                    "Partner tạo khách sạn hợp lệ",
                    "Partner đã đăng nhập và có quyền partner.",
                    [
                        "Mở form tạo khách sạn.",
                        "Nhập tên, loại hình, sao, giờ check-in/out, địa chỉ đầy đủ và tiện nghi.",
                        "Nhấn lưu.",
                    ],
                    [
                        "API tạo hotel với ownerId là partner hiện tại.",
                        "Status mặc định là draft.",
                        "Địa chỉ và tiện nghi được lưu đúng.",
                    ],
                ),
                (
                    "Từ chối tên khách sạn quá ngắn",
                    "Partner đang ở form khách sạn.",
                    [
                        "Nhập tên khách sạn chỉ 1 ký tự.",
                        "Nhập các trường còn lại hợp lệ.",
                        "Nhấn lưu.",
                    ],
                    [
                        "Form hoặc API báo tên phải có ít nhất 2 ký tự.",
                        "Không tạo hotel mới.",
                        "Người dùng thấy trường cần sửa.",
                    ],
                ),
                (
                    "Từ chối starRating ngoài 1 đến 5",
                    "Partner đang tạo hoặc sửa khách sạn.",
                    [
                        "Gửi starRating = 0 hoặc 6 qua API.",
                        "Quan sát response.",
                        "Kiểm tra dữ liệu hotel.",
                    ],
                    [
                        "API trả lỗi validation.",
                        "Không lưu starRating sai.",
                        "UI chỉ cho chọn 1 đến 5 sao.",
                    ],
                ),
                (
                    "Từ chối địa chỉ quá ngắn",
                    "Partner đang tạo khách sạn.",
                    [
                        "Nhập addressLine dưới 5 ký tự.",
                        "Chọn tỉnh/quận/phường hợp lệ.",
                        "Nhấn lưu.",
                    ],
                    [
                        "Form hiển thị lỗi địa chỉ tối thiểu 5 ký tự.",
                        "Không gọi API thành công.",
                        "Các lựa chọn địa chỉ khác vẫn được giữ.",
                    ],
                ),
                (
                    "Từ chối tọa độ latitude/longitude ngoài biên",
                    "Partner hoặc API client gửi tọa độ.",
                    [
                        "Gửi latitude = 91 hoặc longitude = 181.",
                        "Gọi create/update hotel.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả lỗi validation.",
                        "Không lưu tọa độ ngoài [-90,90] và [-180,180].",
                        "Dữ liệu địa chỉ hiện có không bị ghi đè sai.",
                    ],
                ),
                (
                    "Cập nhật khách sạn thuộc partner hiện tại",
                    "Partner đã có một khách sạn draft/approved.",
                    [
                        "Mở màn hình edit hotel.",
                        "Thay đổi mô tả, tiện nghi hoặc giờ nhận/trả phòng.",
                        "Nhấn lưu.",
                    ],
                    [
                        "API cập nhật đúng hotel thuộc owner hiện tại.",
                        "Thông tin mới hiển thị sau khi reload.",
                        "Không thay đổi ownerId hoặc status trái phép.",
                    ],
                ),
                (
                    "Không cho Partner sửa khách sạn của partner khác",
                    "Tồn tại hotel thuộc partner B.",
                    [
                        "Đăng nhập partner A.",
                        "Gọi PUT /api/v1/partner/hotels/:id với id của partner B.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả 404 hoặc 403.",
                        "Không lộ dữ liệu chi tiết của hotel thuộc partner khác.",
                        "Không cập nhật record trong database.",
                    ],
                ),
                (
                    "Submit khách sạn để duyệt",
                    "Hotel thuộc partner đang ở status draft và có thông tin tối thiểu.",
                    [
                        "Mở chi tiết khách sạn partner.",
                        "Nhấn Submit for review.",
                        "Quan sát trạng thái.",
                    ],
                    [
                        "Status chuyển từ draft sang pending.",
                        "Partner không thể tự chuyển trực tiếp sang approved.",
                        "Admin có thể nhìn thấy khách sạn trong danh sách chờ duyệt.",
                    ],
                ),
                (
                    "Upload nhiều ảnh khách sạn hợp lệ",
                    "Partner đã có hotel thuộc sở hữu và token hợp lệ.",
                    [
                        "Chọn nhiều file ảnh hợp lệ.",
                        "Gửi POST /api/v1/partner/hotels/:id/images.",
                        "Quan sát danh sách ảnh sau upload.",
                    ],
                    [
                        "File được upload qua middleware uploadMultipleImages.",
                        "HotelImage được tạo đúng hotelId.",
                        "UI hiển thị ảnh mới và không mất ảnh cũ.",
                    ],
                ),
                (
                    "Từ chối upload file không phải ảnh ở endpoint ảnh",
                    "Partner đã đăng nhập.",
                    [
                        "Chọn file không phải ảnh, ví dụ .txt hoặc .exe.",
                        "Gửi vào endpoint upload images.",
                        "Quan sát response.",
                    ],
                    [
                        "Middleware từ chối file không hợp lệ.",
                        "Không tạo HotelImage.",
                        "UI hiển thị lỗi upload rõ ràng.",
                    ],
                ),
                (
                    "Xóa ảnh khách sạn sau xác nhận",
                    "Hotel có ít nhất một ảnh và partner là owner.",
                    [
                        "Nhấn nút xóa ảnh.",
                        "Xác nhận trong modal.",
                        "Quan sát danh sách ảnh.",
                    ],
                    [
                        "Ảnh bị xóa khỏi hotel hiện tại.",
                        "Modal đóng sau khi xóa thành công.",
                        "Không xóa ảnh khác ngoài imageId đã chọn.",
                    ],
                ),
                (
                    "Xóa khách sạn draft",
                    "Partner có hotel draft không còn cần sử dụng.",
                    [
                        "Mở danh sách khách sạn partner.",
                        "Chọn xóa hotel draft.",
                        "Xác nhận thao tác.",
                    ],
                    [
                        "Hotel bị xóa hoặc soft-delete theo service hiện tại.",
                        "Danh sách không còn hiển thị hotel đã xóa.",
                        "Các record liên quan được xử lý đúng theo quan hệ cascade hoặc policy.",
                    ],
                ),
            ],
        ),
        (
            "PARTNER_ROOM",
            "Partner - Phòng, Giá & Inventory",
            [
                (
                    "Tạo loại phòng hợp lệ với đủ bảng giá",
                    "Partner có khách sạn thuộc sở hữu.",
                    [
                        "Mở form thêm loại phòng.",
                        "Nhập tên, sức chứa, loại giường, diện tích, tổng số phòng.",
                        "Nhập giá theo giờ, qua đêm và theo ngày rồi lưu.",
                    ],
                    [
                        "RoomType được tạo với hotelId đúng.",
                        "Các pricingPolicies hourly, overnight, daily được tạo nếu có nhập giá.",
                        "UI hiển thị thông báo thêm loại phòng thành công.",
                    ],
                ),
                (
                    "Từ chối tên loại phòng rỗng",
                    "Partner đang ở form RoomForm.",
                    [
                        "Để trống tên loại phòng.",
                        "Nhập các trường giá hợp lệ.",
                        "Nhấn lưu.",
                    ],
                    [
                        "Form báo vui lòng nhập tên loại phòng.",
                        "Không tạo room type.",
                        "Focus hoặc lỗi nằm gần trường tên.",
                    ],
                ),
                (
                    "Từ chối maxGuests nhỏ hơn 1",
                    "Partner đang tạo phòng.",
                    [
                        "Nhập maxGuests = 0.",
                        "Nhấn lưu.",
                        "Quan sát form/API.",
                    ],
                    [
                        "Validation báo sức chứa tối thiểu 1 người.",
                        "Không lưu dữ liệu sai.",
                        "UI không crash khi nhập số 0.",
                    ],
                ),
                (
                    "Từ chối maxGuests vượt quá 20 qua API",
                    "Partner có token hợp lệ.",
                    [
                        "Gửi POST room-types với maxGuests = 21.",
                        "Quan sát response.",
                        "Kiểm tra database.",
                    ],
                    [
                        "API trả lỗi validation max 20.",
                        "Không tạo RoomType.",
                        "Thông báo lỗi có thể map lên UI.",
                    ],
                ),
                (
                    "Từ chối totalUnits nhỏ hơn 1",
                    "Partner đang tạo hoặc sửa room type.",
                    [
                        "Nhập totalUnits = 0.",
                        "Nhấn lưu.",
                        "Quan sát response.",
                    ],
                    [
                        "Validation báo tối thiểu 1 phòng.",
                        "Không lưu totalUnits sai.",
                        "Inventory không được tạo với số phòng âm hoặc 0.",
                    ],
                ),
                (
                    "Tạo room unit hợp lệ",
                    "RoomType thuộc hotel của partner đã tồn tại.",
                    [
                        "Gọi POST /units với roomNumber hợp lệ.",
                        "Nhập floor trong khoảng -5 đến 200 nếu có.",
                        "Quan sát response.",
                    ],
                    [
                        "RoomUnit được tạo với status mặc định available.",
                        "Không trùng roomNumber trong cùng roomType.",
                        "Danh sách units hiển thị phòng mới.",
                    ],
                ),
                (
                    "Từ chối roomNumber trùng trong cùng room type",
                    "Đã có RoomUnit với roomNumber cụ thể.",
                    [
                        "Tạo thêm RoomUnit cùng roomNumber trong cùng roomType.",
                        "Quan sát response API.",
                        "Kiểm tra danh sách units.",
                    ],
                    [
                        "API trả lỗi unique hoặc thông báo phòng đã tồn tại.",
                        "Không tạo bản ghi trùng.",
                        "Danh sách units không tăng sai.",
                    ],
                ),
                (
                    "Cập nhật status room unit",
                    "RoomUnit thuộc roomType của partner tồn tại.",
                    [
                        "Gửi PUT unit với status maintenance.",
                        "Tải lại danh sách units.",
                        "Đổi status về available.",
                    ],
                    [
                        "Status cập nhật đúng một trong available, occupied, maintenance, cleaning.",
                        "UI badge trạng thái đổi tương ứng.",
                        "Status không thuộc enum bị từ chối.",
                    ],
                ),
                (
                    "Tạo pricing hourly hợp lệ",
                    "RoomType thuộc partner tồn tại.",
                    [
                        "Gửi POST pricing với bookingType hourly, basePrice dương, minHours hợp lệ.",
                        "Quan sát response.",
                        "Tải lại pricing list.",
                    ],
                    [
                        "PricingPolicy hourly được tạo.",
                        "basePrice là số dương.",
                        "Không tạo trùng bookingType cho cùng roomType nếu đã tồn tại.",
                    ],
                ),
                (
                    "Từ chối giá basePrice không dương",
                    "Partner có roomType hợp lệ.",
                    [
                        "Gửi pricing với basePrice = 0 hoặc âm.",
                        "Quan sát response.",
                        "Kiểm tra dữ liệu pricing.",
                    ],
                    [
                        "API trả lỗi giá phải lớn hơn 0.",
                        "Không tạo hoặc cập nhật pricing sai.",
                        "UI hiển thị lỗi nhập giá.",
                    ],
                ),
                (
                    "Tạo special price hợp lệ",
                    "PricingPolicy tồn tại.",
                    [
                        "Gửi special price với date YYYY-MM-DD và price dương.",
                        "Tải danh sách special prices.",
                        "Kiểm tra giá theo ngày.",
                    ],
                    [
                        "SpecialPrice được tạo đúng pricingPolicyId.",
                        "Không cho trùng date trong cùng pricingPolicy.",
                        "reason được lưu nếu có.",
                    ],
                ),
                (
                    "Cập nhật inventory theo room type",
                    "Partner có hotel và roomType.",
                    [
                        "Mở màn hình inventory calendar.",
                        "Cập nhật totalRooms/bookedRooms/isClosed cho một ngày.",
                        "Tải lại calendar.",
                    ],
                    [
                        "Inventory của đúng roomType/date được cập nhật.",
                        "Không cho bookedRooms vượt totalRooms nếu service có rule.",
                        "Ngày đóng phòng hiển thị không khả dụng với khách.",
                    ],
                ),
            ],
        ),
        (
            "PARTNER_OPS",
            "Partner - Booking, Thống kê & Cài đặt",
            [
                (
                    "Partner xem booking thuộc khách sạn của mình",
                    "Partner có khách sạn đã phát sinh booking.",
                    [
                        "Đăng nhập partner.",
                        "Mở màn hình Booking Management.",
                        "Gọi GET /api/v1/partner/bookings.",
                    ],
                    [
                        "Danh sách chỉ gồm booking thuộc hotel của partner hiện tại.",
                        "Có thông tin khách, phòng, thời gian, giá và trạng thái.",
                        "Không hiển thị booking của partner khác.",
                    ],
                ),
                (
                    "Lọc booking partner theo trạng thái",
                    "Partner có booking ở nhiều trạng thái.",
                    [
                        "Chọn filter CONFIRMED hoặc CANCELLED.",
                        "Gọi API với query status.",
                        "Quan sát danh sách.",
                    ],
                    [
                        "Danh sách chỉ hiển thị booking đúng trạng thái filter.",
                        "Chọn ALL trả lại toàn bộ booking thuộc partner.",
                        "UI không mất filter khi reload dữ liệu.",
                    ],
                ),
                (
                    "Partner cập nhật trạng thái booking hợp lệ",
                    "Booking thuộc khách sạn của partner đang ở trạng thái có thể cập nhật.",
                    [
                        "Mở booking detail hoặc danh sách.",
                        "Chọn trạng thái mới hợp lệ.",
                        "Gửi PATCH /api/v1/partner/bookings/:id/status.",
                    ],
                    [
                        "Booking được cập nhật trạng thái mới.",
                        "Response trả BOOKING_STATUS_UPDATED.",
                        "Customer nhìn thấy trạng thái mới trong danh sách booking.",
                    ],
                ),
                (
                    "Không cho partner cập nhật booking của partner khác",
                    "Có booking thuộc hotel của partner B.",
                    [
                        "Đăng nhập partner A.",
                        "Gọi PATCH status booking của partner B.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả 404 hoặc 403.",
                        "Booking không bị thay đổi.",
                        "Không lộ thông tin khách hàng của partner B.",
                    ],
                ),
                (
                    "Từ chối trạng thái booking không hợp lệ",
                    "Partner có booking thuộc quyền.",
                    [
                        "Gửi PATCH status với giá trị ngoài enum.",
                        "Quan sát response.",
                        "Kiểm tra booking.",
                    ],
                    [
                        "API không cập nhật trạng thái sai.",
                        "Trả lỗi validation hoặc lỗi xử lý rõ ràng.",
                        "UI không có option trạng thái ngoài enum.",
                    ],
                ),
                (
                    "Hiển thị thống kê tổng quan partner",
                    "Partner có dữ liệu hotel, room, booking và doanh thu.",
                    [
                        "Mở partner dashboard/stats.",
                        "Quan sát các chỉ số.",
                        "So sánh với dữ liệu booking hiện có.",
                    ],
                    [
                        "Chỉ số tổng booking, doanh thu, phòng và khách sạn thuộc partner hiện tại.",
                        "Không tính dữ liệu của partner khác.",
                        "Biểu đồ không lỗi khi dữ liệu rỗng.",
                    ],
                ),
                (
                    "Cài đặt chính sách hủy phòng",
                    "Partner có hotel thuộc sở hữu.",
                    [
                        "Mở màn hình cancellation policy.",
                        "Chọn flexible/moderate/strict/non_refundable và giờ hủy.",
                        "Lưu thay đổi.",
                    ],
                    [
                        "Hotel cancellationPolicy và cancellationHours cập nhật đúng.",
                        "cancellationHours không âm.",
                        "Thông tin policy hiển thị lại đúng sau reload.",
                    ],
                ),
                (
                    "Cài đặt đặt cọc hợp lệ",
                    "Partner có hotel thuộc sở hữu.",
                    [
                        "Mở màn hình deposit policy.",
                        "Nhập depositPercent trong khoảng 0 đến 100.",
                        "Lưu thay đổi.",
                    ],
                    [
                        "depositPercent được lưu đúng.",
                        "UI hiển thị phần trăm đặt cọc mới.",
                        "Booking sau đó áp dụng thông tin đặt cọc theo nghiệp vụ nếu có.",
                    ],
                ),
                (
                    "Từ chối depositPercent ngoài 0 đến 100",
                    "Partner đang cập nhật policy.",
                    [
                        "Gửi depositPercent = -1 hoặc 101.",
                        "Quan sát response.",
                        "Kiểm tra dữ liệu hotel.",
                    ],
                    [
                        "API trả lỗi validation.",
                        "Không lưu depositPercent sai.",
                        "UI cảnh báo giá trị không hợp lệ.",
                    ],
                ),
                (
                    "Partner settings hiển thị thông tin tài khoản",
                    "Partner đã đăng nhập.",
                    [
                        "Mở màn hình Settings.",
                        "Quan sát thông tin user/hotel nếu có.",
                        "Thử quay lại dashboard.",
                    ],
                    [
                        "Thông tin partner hiển thị đúng.",
                        "Không hiển thị thông tin nhạy cảm như password/token.",
                        "Điều hướng trong khu vực partner ổn định.",
                    ],
                ),
                (
                    "Sidebar partner điều hướng đúng các trang",
                    "Partner đã đăng nhập trên web hoặc mobile layout hỗ trợ.",
                    [
                        "Nhấn lần lượt Dashboard, Booking, Rooms, Vouchers, Settings.",
                        "Quan sát URL và nội dung trang.",
                        "Quay lại trang trước.",
                    ],
                    [
                        "Mỗi menu mở đúng route.",
                        "Active state của sidebar đúng trang hiện tại.",
                        "Không reload mất phiên đăng nhập.",
                    ],
                ),
                (
                    "Thông báo lỗi API partner được hiển thị thân thiện",
                    "Giả lập API partner trả lỗi mạng hoặc lỗi 500.",
                    [
                        "Mở một màn hình partner có gọi API.",
                        "Ngắt backend hoặc mock lỗi.",
                        "Quan sát UI.",
                    ],
                    [
                        "UI hiển thị message modal/error box phù hợp.",
                        "Không crash toàn ứng dụng.",
                        "Người dùng có thể thử lại hoặc quay lại.",
                    ],
                ),
            ],
        ),
        (
            "ADMIN",
            "Admin - Quản trị & Phân quyền",
            [
                (
                    "Admin dashboard chỉ truy cập đầy đủ trên web",
                    "Người dùng admin đăng nhập trên mobile và web.",
                    [
                        "Đăng nhập admin trên mobile/emulator.",
                        "Quan sát màn hình admin.",
                        "Đăng nhập admin trên web desktop.",
                    ],
                    [
                        "Mobile hiển thị cảnh báo yêu cầu truy cập trên web.",
                        "Web hiển thị AdminShell và dashboard.",
                        "Nút quay lại đăng nhập/logout hoạt động trên mobile.",
                    ],
                ),
                (
                    "Sidebar admin hiển thị theo quyền",
                    "Có user admin/operator/accountant với permission khác nhau.",
                    [
                        "Đăng nhập bằng role có quyền hạn chế.",
                        "Quan sát menu sidebar.",
                        "So sánh với permission từ API.",
                    ],
                    [
                        "Chỉ tab có quyền view được hiển thị.",
                        "Nếu activeTab mất quyền, hệ thống chuyển về overview.",
                        "Không thể truy cập tab bị ẩn bằng thao tác UI.",
                    ],
                ),
                (
                    "Global search trong admin điều hướng đúng",
                    "Admin đang ở dashboard web.",
                    [
                        "Nhập keyword tên menu vào ô global search.",
                        "Chọn một kết quả.",
                        "Quan sát active tab.",
                    ],
                    [
                        "Search trả tối đa 8 kết quả phù hợp.",
                        "Nhấn kết quả sẽ điều hướng đúng tab.",
                        "Ô search được reset sau khi điều hướng.",
                    ],
                ),
                (
                    "Admin xem danh sách người dùng có phân trang",
                    "Có nhiều user trong hệ thống.",
                    [
                        "Mở tab quản lý người dùng.",
                        "Gọi API GET /api/admin/users với page và limit.",
                        "Chuyển trang.",
                    ],
                    [
                        "Danh sách người dùng hiển thị đúng page.",
                        "Tổng số bản ghi và pagination đúng.",
                        "Không hiển thị password/refreshToken.",
                    ],
                ),
                (
                    "Admin tìm kiếm user theo email hoặc username",
                    "Có user khớp từ khóa.",
                    [
                        "Nhập từ khóa vào ô search user.",
                        "Quan sát request và kết quả.",
                        "Xóa từ khóa.",
                    ],
                    [
                        "Kết quả được lọc theo từ khóa.",
                        "Page reset về 1 khi đổi từ khóa.",
                        "Xóa từ khóa trả lại danh sách ban đầu.",
                    ],
                ),
                (
                    "Admin tạo tài khoản nhân viên hợp lệ",
                    "Admin có quyền users create/update theo UI.",
                    [
                        "Mở modal tạo user.",
                        "Nhập username, email, password và role OPERATOR hoặc ACCOUNTANT.",
                        "Nhấn tạo.",
                    ],
                    [
                        "User mới được tạo thành công.",
                        "Modal đóng và form reset.",
                        "Danh sách user reload có tài khoản mới.",
                    ],
                ),
                (
                    "Không cho tạo user khi thiếu trường bắt buộc",
                    "Admin đang ở modal tạo user.",
                    [
                        "Để trống username hoặc email hoặc password.",
                        "Nhấn tạo.",
                        "Quan sát alert.",
                    ],
                    [
                        "UI báo vui lòng điền đầy đủ thông tin.",
                        "Không gọi API tạo user thành công.",
                        "Modal vẫn mở để nhập lại.",
                    ],
                ),
                (
                    "Không cho admin tự khóa tài khoản của chính mình",
                    "Admin đang đăng nhập và nhìn thấy tài khoản của mình trong danh sách.",
                    [
                        "Nhấn khóa tài khoản của chính user hiện tại.",
                        "Quan sát alert.",
                        "Kiểm tra status user.",
                    ],
                    [
                        "UI báo không thể tự khóa tài khoản chính mình.",
                        "Không gọi API block hoặc API trả lỗi.",
                        "Tài khoản hiện tại vẫn active.",
                    ],
                ),
                (
                    "Chỉ Super Admin quản lý tài khoản Super Admin",
                    "Có user SUPER_ADMIN và một admin thường.",
                    [
                        "Đăng nhập admin thường.",
                        "Thử khóa, xóa hoặc đổi role SUPER_ADMIN.",
                        "Quan sát response/UI.",
                    ],
                    [
                        "Admin thường bị từ chối thao tác với SUPER_ADMIN.",
                        "Chỉ SUPER_ADMIN được phép theo policy.",
                        "Không thay đổi dữ liệu tài khoản được bảo vệ.",
                    ],
                ),
                (
                    "Cập nhật phân quyền role và cảnh báo thay đổi chưa lưu",
                    "Admin có quyền roles update hoặc SUPER_ADMIN.",
                    [
                        "Mở tab Phân quyền.",
                        "Thay đổi một quyền nhưng chưa lưu.",
                        "Chuyển sang tab khác.",
                    ],
                    [
                        "UI hiển thị cảnh báo có thay đổi chưa lưu trên web.",
                        "Nếu người dùng hủy, vẫn ở tab phân quyền.",
                        "Nếu đồng ý rời trang, trạng thái dirty được reset.",
                    ],
                ),
                (
                    "Admin duyệt hoặc từ chối cơ sở lưu trú",
                    "Có hotel/property đang pending.",
                    [
                        "Mở tab Lodging.",
                        "Chọn một cơ sở pending.",
                        "Cập nhật status approved hoặc rejected.",
                    ],
                    [
                        "Status cơ sở lưu trú được cập nhật đúng.",
                        "approvedAt/approvedBy hoặc rejectionReason được xử lý theo service.",
                        "Customer chỉ nhìn thấy cơ sở đã được duyệt nếu policy yêu cầu.",
                    ],
                ),
                (
                    "Admin quản lý review",
                    "Có review ở trạng thái PENDING.",
                    [
                        "Mở tab Reviews.",
                        "Chọn approve hoặc hide review.",
                        "Quan sát danh sách sau cập nhật.",
                    ],
                    [
                        "Review chuyển trạng thái đúng.",
                        "Danh sách reload hoặc cập nhật tại chỗ.",
                        "Không cho user thiếu quyền approve cập nhật review.",
                    ],
                ),
                (
                    "Admin quản lý content",
                    "Admin có quyền content create/update/delete.",
                    [
                        "Mở tab Content.",
                        "Tạo bài viết với title, category và body.",
                        "Cập nhật status published hoặc archived.",
                    ],
                    [
                        "ContentPost được tạo/cập nhật đúng.",
                        "Trạng thái content nằm trong DRAFT/PUBLISHED/ARCHIVED.",
                        "Danh sách content phản ánh thay đổi.",
                    ],
                ),
                (
                    "Admin export dữ liệu hợp lệ",
                    "Admin có quyền export resource tương ứng.",
                    [
                        "Tại tab Users hoặc Bookings, nhấn Export.",
                        "Gọi GET /api/admin/export/:resource.",
                        "Quan sát file tải về.",
                    ],
                    [
                        "API trả file export đúng resource.",
                        "Nếu resource không hỗ trợ, API trả 400 EXPORT_RESOURCE_UNSUPPORTED.",
                        "Nếu thiếu quyền export, API trả 403.",
                    ],
                ),
            ],
        ),
        (
            "NONFUNC",
            "Bảo mật, UI/UX & Hiệu năng",
            [
                (
                    "API bảo vệ trả 401 khi thiếu token",
                    "Người dùng không gửi Authorization header.",
                    [
                        "Gọi endpoint cần đăng nhập như /api/customer/bookings.",
                        "Gọi endpoint /api/v1/partner/hotels.",
                        "Quan sát response.",
                    ],
                    [
                        "Các endpoint bảo vệ trả 401.",
                        "Response không chứa stack trace hoặc dữ liệu nội bộ.",
                        "Header/format lỗi nhất quán.",
                    ],
                ),
                (
                    "API role guard trả 403 khi sai role",
                    "Customer có token hợp lệ nhưng không phải partner/admin.",
                    [
                        "Dùng token customer gọi endpoint partner.",
                        "Dùng token partner gọi endpoint admin.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả 403 Forbidden.",
                        "Không thực thi logic nghiệp vụ sau guard.",
                        "Không rò rỉ dữ liệu resource.",
                    ],
                ),
                (
                    "Không chấp nhận role admin khi đăng ký public",
                    "Người dùng gọi API register public.",
                    [
                        "Gửi payload đăng ký với role = admin.",
                        "Quan sát response.",
                        "Kiểm tra bảng users.",
                    ],
                    [
                        "API từ chối role ngoài customer/partner.",
                        "Không tạo tài khoản admin qua public register.",
                        "Thông báo lỗi rõ ràng.",
                    ],
                ),
                (
                    "Chống truy cập chéo dữ liệu partner",
                    "Có dữ liệu khách sạn/phòng/booking thuộc nhiều partner.",
                    [
                        "Đăng nhập partner A.",
                        "Thử đọc/sửa/xóa resource của partner B.",
                        "Kiểm tra log và database.",
                    ],
                    [
                        "Backend kiểm tra ownerId/hotel ownership ở service.",
                        "Request bị từ chối.",
                        "Dữ liệu partner B không bị thay đổi hoặc rò rỉ.",
                    ],
                ),
                (
                    "Chống truy cập chéo dữ liệu customer",
                    "Có booking/profile/message của nhiều customer.",
                    [
                        "Đăng nhập customer A.",
                        "Thử đọc booking hoặc message của customer B bằng id trực tiếp.",
                        "Quan sát response.",
                    ],
                    [
                        "API trả 403 hoặc 404.",
                        "Không rò rỉ tên, số điện thoại, mã thanh toán hoặc thông tin khách sạn riêng.",
                        "Không ghi nhận viewed/notification cho sai user.",
                    ],
                ),
                (
                    "Validation chống payload quá dài",
                    "Client có thể gửi payload trực tiếp qua API.",
                    [
                        "Gửi username, hotel description hoặc room description vượt max length.",
                        "Quan sát response.",
                        "Kiểm tra dữ liệu lưu.",
                    ],
                    [
                        "API trả lỗi validation.",
                        "Không lưu chuỗi vượt giới hạn.",
                        "Response không làm server crash.",
                    ],
                ),
                (
                    "Xử lý ký tự đặc biệt trong tìm kiếm và form",
                    "Có quyền nhập text ở search, hotel name, content hoặc support.",
                    [
                        "Nhập chuỗi chứa dấu nháy, dấu phần trăm, script tag và tiếng Việt có dấu.",
                        "Submit hoặc tìm kiếm.",
                        "Quan sát UI và response.",
                    ],
                    [
                        "Ứng dụng không thực thi script người dùng nhập.",
                        "Text được hiển thị an toàn hoặc encode đúng.",
                        "Tìm kiếm không lỗi SQL/Prisma.",
                    ],
                ),
                (
                    "Giới hạn kích thước JSON request",
                    "Backend Express cấu hình json limit 10mb.",
                    [
                        "Gửi JSON request nhỏ hơn 10mb tới endpoint hợp lệ.",
                        "Gửi JSON vượt 10mb.",
                        "Quan sát response.",
                    ],
                    [
                        "Payload hợp lệ được xử lý bình thường.",
                        "Payload quá lớn bị từ chối.",
                        "Server vẫn hoạt động cho request tiếp theo.",
                    ],
                ),
                (
                    "Upload file lớn hoặc sai MIME không làm sập server",
                    "Endpoint upload file yêu cầu token hợp lệ.",
                    [
                        "Upload file vượt giới hạn hoặc MIME không hợp lệ.",
                        "Quan sát response.",
                        "Gửi request upload hợp lệ sau đó.",
                    ],
                    [
                        "Request sai bị từ chối có thông báo rõ ràng.",
                        "Không lưu file không hợp lệ.",
                        "Server tiếp tục xử lý file hợp lệ sau lỗi.",
                    ],
                ),
                (
                    "Hiệu năng danh sách khách sạn limit 50",
                    "Database có đủ dữ liệu hotel_cards.",
                    [
                        "Gọi GET /api/customer/hotels?limit=50.",
                        "Đo thời gian phản hồi trong môi trường test.",
                        "Lặp lại với filter và sort.",
                    ],
                    [
                        "API trả phản hồi trong ngưỡng chấp nhận của dự án.",
                        "Không timeout khi sort/filter phổ biến.",
                        "Response size phù hợp cho mobile.",
                    ],
                ),
                (
                    "Hiệu năng kiểm tra availability nhiều slot",
                    "Khách sạn có nhiều room types và booking.",
                    [
                        "Gọi availability Theo giờ cho ngày tương lai.",
                        "Đo thời gian phản hồi.",
                        "Kiểm tra số lượng query nếu có logging.",
                    ],
                    [
                        "API trả đủ slot 30 phút.",
                        "Không timeout khi dữ liệu booking tăng.",
                        "Kết quả maxHours nhất quán giữa các lần gọi.",
                    ],
                ),
                (
                    "UI không vỡ layout trên màn hình nhỏ",
                    "Ứng dụng chạy ở viewport mobile hẹp.",
                    [
                        "Mở Login, Register, Hotel Detail, Booking Payment, Partner Room Form.",
                        "Cuộn toàn bộ màn hình.",
                        "Quan sát text, nút, ảnh và form.",
                    ],
                    [
                        "Không có text chồng lấn hoặc bị cắt khó đọc.",
                        "Các nút chính nằm trong vùng thao tác được.",
                        "Input dài được wrap hoặc scroll hợp lý.",
                    ],
                ),
                (
                    "UI admin web không vỡ khi sidebar thu gọn",
                    "Admin đăng nhập trên web desktop.",
                    [
                        "Nhấn nút collapse sidebar.",
                        "Mở từng menu chính và menu con.",
                        "Bật/tắt dark mode.",
                    ],
                    [
                        "Sidebar thu gọn vẫn hiển thị icon và active state.",
                        "Nội dung chính không bị che.",
                        "Dark mode giữ tương phản và không làm mất text.",
                    ],
                ),
                (
                    "Thông báo lỗi backend không lộ stack trace",
                    "Gây lỗi validation, lỗi không có quyền và lỗi resource not found.",
                    [
                        "Gửi request sai tới các endpoint đại diện.",
                        "Quan sát response body.",
                        "Kiểm tra console/log server nếu có.",
                    ],
                    [
                        "Client chỉ nhận message và code phù hợp.",
                        "Stack trace chỉ nằm ở log server, không trả cho client.",
                        "Format lỗi nhất quán giữa customer, partner và admin.",
                    ],
                ),
            ],
        ),
    ]

    cases: list[dict[str, str]] = []
    counters: dict[str, int] = {}
    for prefix, module, specs in groups:
        counters[prefix] = 0
        for title, preconditions, steps, expected in specs:
            counters[prefix] += 1
            cases.append(
                {
                    "Mã TC": f"TC_{prefix}_{counters[prefix]:02d}",
                    "Phân Hệ": module,
                    "Tên Kiểm Thử": title,
                    "Điều Kiện Tiên Quyết": preconditions,
                    "Các Bước Thực Hiện": numbered(steps),
                    "Kết Quả Mong Đợi": numbered(expected),
                    "Trạng Thái": "Not Run",
                    "Ghi Chú": "",
                }
            )

    return cases


def apply_common_sheet_settings(ws) -> None:
    ws.sheet_view.showGridLines = True


def style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(name="Calibri", color=WHITE, bold=True, size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def apply_table_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def create_detail_sheet(wb: Workbook, test_cases: list[dict[str, str]]) -> None:
    ws = wb.create_sheet(DETAIL_SHEET)
    apply_common_sheet_settings(ws)

    headers = [
        "Mã TC",
        "Phân Hệ",
        "Tên Kiểm Thử",
        "Điều Kiện Tiên Quyết",
        "Các Bước Thực Hiện",
        "Kết Quả Mong Đợi",
        "Trạng Thái",
        "Ghi Chú",
    ]
    ws.append(headers)
    for case in test_cases:
        ws.append([case[header] for header in headers])

    style_header_row(ws, 1, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    widths = {
        "A": 15,
        "B": 26,
        "C": 40,
        "D": 42,
        "E": 52,
        "F": 54,
        "G": 15,
        "H": 24,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    center_columns = {"A", "B", "G"}
    left_columns = {"C", "D", "E", "F", "H"}
    for row_idx in range(2, ws.max_row + 1):
        fill_color = ZEBRA if row_idx % 2 == 1 else WHITE
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_letter = get_column_letter(col_idx)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Calibri", size=10, color=TEXT_DARK)
            cell.alignment = Alignment(
                horizontal="center" if col_letter in center_columns else "left",
                vertical="top",
                wrap_text=True,
            )
            if col_letter in left_columns:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 92

    ws.row_dimensions[1].height = 30
    apply_table_borders(ws, 1, ws.max_row, 1, len(headers))

    status_validation = DataValidation(type="list", formula1='"Not Run,Pass,Fail"', allow_blank=False)
    status_validation.error = "Chỉ được chọn Not Run, Pass hoặc Fail."
    status_validation.errorTitle = "Trạng thái không hợp lệ"
    status_validation.prompt = "Chọn trạng thái chạy test case."
    status_validation.promptTitle = "Trạng thái"
    ws.add_data_validation(status_validation)
    status_validation.add(f"G2:G{ws.max_row}")


def create_summary_sheet(wb: Workbook, test_cases: list[dict[str, str]]) -> None:
    ws = wb[SUMMARY_SHEET]
    apply_common_sheet_settings(ws)

    modules = []
    for case in test_cases:
        module = case["Phân Hệ"]
        if module not in modules:
            modules.append(module)

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "BÁO CÁO VÀ KỊCH BẢN KIỂM THỬ PHẦN MỀM"
    title.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    metadata = [
        ("Người thực hiện", "Senior QA"),
        ("Ngày tạo", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Trạng thái", "Ready for Execution"),
    ]
    for row_idx, (label, value) in enumerate(metadata, 3):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True, color=TEXT_DARK)
        ws.cell(row=row_idx, column=2).font = Font(name="Calibri", color=TEXT_DARK)
        ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(row=row_idx, column=2).fill = PatternFill("solid", fgColor=WHITE)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center")

    header_row = 7
    headers = ["STT", "Phân Hệ Kiểm Thử (Module)", "Tổng số TC", "Đạt (Pass)", "Lỗi (Fail)", "Chưa chạy (Not Run)"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=header)
    style_header_row(ws, header_row, 1, len(headers))

    first_data_row = header_row + 1
    detail_ref = f"'{DETAIL_SHEET}'"
    for offset, module in enumerate(modules):
        row = first_data_row + offset
        ws.cell(row=row, column=1, value=offset + 1)
        ws.cell(row=row, column=2, value=module)
        ws.cell(row=row, column=3, value=f'=COUNTIF({detail_ref}!$B:$B,B{row})')
        ws.cell(row=row, column=4, value=f'=COUNTIFS({detail_ref}!$B:$B,B{row},{detail_ref}!$G:$G,"Pass")')
        ws.cell(row=row, column=5, value=f'=COUNTIFS({detail_ref}!$B:$B,B{row},{detail_ref}!$G:$G,"Fail")')
        ws.cell(row=row, column=6, value=f"=C{row}-D{row}-E{row}")

    total_row = first_data_row + len(modules)
    ws.cell(row=total_row, column=1, value="Tổng")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    for col_idx in range(3, 7):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}{first_data_row}:{col_letter}{total_row - 1})")

    for row_idx in range(first_data_row, total_row + 1):
        fill_color = LIGHT_GREEN if row_idx == total_row else (ZEBRA if row_idx % 2 == 1 else WHITE)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Calibri", size=10, bold=row_idx == total_row, color=TEXT_DARK)
            cell.alignment = Alignment(
                horizontal="center" if col_idx != 2 else "left",
                vertical="center",
                wrap_text=True,
            )

    widths = {
        "A": 10,
        "B": 34,
        "C": 15,
        "D": 15,
        "E": 15,
        "F": 20,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    apply_table_borders(ws, 3, 5, 1, 2)
    apply_table_borders(ws, header_row, total_row, 1, len(headers))
    ws.freeze_panes = "A8"


def save_workbook() -> Path:
    test_cases = build_test_cases()

    wb = Workbook()
    ws = wb.active
    ws.title = SUMMARY_SHEET

    create_summary_sheet(wb, test_cases)
    create_detail_sheet(wb, test_cases)

    output_path = Path(__file__).resolve().parent / OUTPUT_FILE
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = save_workbook()
    print(f"Created: {path}")
